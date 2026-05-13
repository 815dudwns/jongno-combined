"""
AMI 작업 데이터(신규 스키마) → 도로명주소 + 좌표 동시 추출 (병렬)
대상리스트_260428.xlsx → ami_data_coords.json
"""

import openpyxl
import json
import requests
import re
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

API_KEY     = 'e46ada1811d067b4acf77d992a13b52e'
EXCEL_FILE  = '대상리스트_260428.xlsx'
OUTPUT_JSON = 'ami_data_coords.json'
WORKERS     = 10


def parse_type(meter_no):
    code = str(meter_no)[2:4]
    if code == '17': return 'E'
    if code == '19': return 'EA'
    if code in ('25', '26', '27', '45', '46', '47'): return 'G'
    if code in ('53', '55'): return 'Amigo'
    return '알수없음'


def fix_meter_no(val):
    if val is None: return ''.zfill(11)
    s = re.sub(r'[가-힣\s]', '', str(val))  # 한글 prefix("상계" 등) 제거
    try: return str(int(float(s))).zfill(11)
    except (ValueError, TypeError): return s.zfill(11)


_session = requests.Session()
_session.headers.update({'Authorization': f'KakaoAK {API_KEY}'})

def search_address(query):
    if not query: return None
    import time as _t
    for attempt in range(4):
        try:
            r = _session.get('https://dapi.kakao.com/v2/local/search/address.json',
                             params={'query': query}, timeout=15)
            if r.status_code == 429:
                _t.sleep(0.5 * (attempt + 1))
                continue
            if r.status_code == 200:
                docs = r.json().get('documents', [])
                if docs:
                    doc = docs[0]
                    lat = float(doc['y']); lng = float(doc['x'])
                    road = (doc.get('road_address') or {}).get('address_name', '')
                    jibun = (doc.get('address') or {}).get('address_name', '')
                    return (road or jibun or query), lat, lng
                return None
            _t.sleep(0.3)
        except Exception:
            _t.sleep(0.3)
    return None


def search_keyword(query):
    """주소검색 실패 시 키워드(장소) 검색으로 재시도 — 시설명·동 단위에 유효"""
    if not query: return None
    try:
        r = _session.get('https://dapi.kakao.com/v2/local/search/keyword.json',
                         params={'query': query, 'size': 1}, timeout=15)
        if r.status_code == 200:
            docs = r.json().get('documents', [])
            if docs:
                doc = docs[0]
                return doc.get('road_address_name') or doc.get('address_name') or query, float(doc['y']), float(doc['x'])
    except Exception:
        pass
    return None


def extract_dong(address):
    m = re.match(r'(.*?[동읍면])', address or '')
    return m.group(1).strip() if m else None


def trim_addr(addr):
    """건물번호 뒤 부속(층/호) 등을 제거 + '길' 단위 끊기"""
    if not addr: return ''
    a = re.sub(r'\s+', ' ', addr).strip()
    # 괄호/대시 뒤 잘라내기
    a = re.split(r'[\(]', a)[0].strip()
    return a


def resolve(key, items):
    jibun = str(items[0].get('지번') or '').strip()
    road = str(items[0].get('도로명') or '').strip()
    jibun_t = trim_addr(jibun)
    road_t = trim_addr(road)

    found = None; accuracy = None
    # 1) 도로명 정확
    if road_t:
        found = search_address(road_t)
        if found: accuracy = 'exact'
    # 2) 지번 정확
    if not found and jibun_t:
        found = search_address(jibun_t)
        if found: accuracy = 'exact'
    # 3) 키워드(장소) 검색 — 도로명/지번 둘 다 실패할 때
    if not found and road_t:
        found = search_keyword(road_t)
        if found: accuracy = 'exact'
    if not found and jibun_t:
        found = search_keyword(jibun_t)
        if found: accuracy = 'exact'
    # 4) 동 폴백
    if not found:
        dong = extract_dong(jibun_t) or extract_dong(road_t)
        if dong:
            found = search_address(dong) or search_keyword(dong)
            if found: accuracy = 'approximate'
    if found:
        return key, accuracy, found[0], found[1], found[2]
    return key, 'fail', road_t or jibun_t, None, None


def main():
    print(f"📍 시작 (워커 {WORKERS}개)", flush=True)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    print(f"총 {len(rows):,}행", flush=True)

    grouped = {}
    for d in rows:
        jibun = str(d.get('지번') or '').strip()
        road = str(d.get('도로명') or '').strip()
        key = f"{jibun}|{road}"
        grouped.setdefault(key, []).append(d)

    total = len(grouped)
    print(f"고유 주소 그룹: {total:,}", flush=True)

    resolved = {}
    counter = {'done': 0, 'exact': 0, 'approx': 0, 'fail': 0}
    lock = threading.Lock()

    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futures = [ex.submit(resolve, k, v) for k, v in grouped.items()]
        for fut in as_completed(futures):
            key, accuracy, road, lat, lng = fut.result()
            resolved[key] = (accuracy, road, lat, lng)
            with lock:
                counter['done'] += 1
                if accuracy == 'exact': counter['exact'] += 1
                elif accuracy == 'approximate': counter['approx'] += 1
                else: counter['fail'] += 1
                if counter['done'] % 200 == 0 or counter['done'] == total:
                    print(f"[{counter['done']:,}/{total:,}] exact={counter['exact']:,} approx={counter['approx']:,} fail={counter['fail']:,}", flush=True)

    result = []
    for key, items in grouped.items():
        accuracy, road, lat, lng = resolved[key]
        for d in items:
            meter_no = fix_meter_no(d.get('계기번호'))
            sangho_val = d.get('공동주택명')
            sangho = str(sangho_val).strip() if sangho_val and sangho_val != 0 and str(sangho_val) != '0' else ''
            jibun_str = str(d.get('지번') or '').strip()
            result.append({
                '지사': d.get('지사') or '',
                '주소': jibun_str,
                '도로명주소': road,
                '계기번호': meter_no,
                '계기타입': parse_type(meter_no),
                '변대주': '',
                '상호': sangho,
                'lat': lat,
                'lng': lng,
                '좌표정확도': accuracy,
                '고객번호': str(d.get('고객번호') or '').strip(),
                'DCUID': str(d.get('DCU ID') or '').strip(),
            })

    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 완료: 입력 {len(rows):,} / 그룹 {total:,} / exact {counter['exact']:,} / approx {counter['approx']:,} / fail {counter['fail']:,} / 계기 {len(result):,}", flush=True)


if __name__ == '__main__':
    main()
