#!/usr/bin/env python3
"""종로 site-data 에 DCU ID 부착.

매칭 원칙(2026-08-12 확정): DCUID 는 DCUID 끼리, 변대주명은 변대주명끼리만 매칭한다.
종로 site-data 의 `변대주` 필드는 이미 전산화번호(8자리, 예 9926G874)라
DCU 대장의 `변대주번호` 컬럼과 직접 매칭한다. 변대주명(한글)은 검증용으로만 쓴다.

DCU ID = 변대주번호(8) + 차수코드(2).  예) 9926G874 -> 9926G87439

사용:
    python3 jongno-combined/scripts/apply_jongno_dcuid.py \
        jongno-combined/data/jongno-site-data.json \
        data/reference/간선망_해지_정지대상.xlsx
"""
import json
import re
import shutil
import sys
from collections import Counter
from datetime import datetime
from zoneinfo import ZoneInfo

import openpyxl

SHEET = '전체DCU 현황'
SHEET_REVIEW = '검토'      # 철거 예정 개소 검토 결과 (검토 컬럼 = 해지/유지)
KST = ZoneInfo('Asia/Seoul')


def unformula(v):
    """엑셀 텍스트 보존 수식(`="0126E043"`) 을 원문자열로 되돌린다."""
    s = '' if v is None else str(v).strip()
    m = re.match(r'^="?(.*?)"?$', s)
    if m:
        s = m.group(1)
    return s.strip()


def load_dcu_ledger(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
    ws = wb[SHEET]
    rows = ws.iter_rows(min_row=2, values_only=True)
    header = [str(x).strip() if x is not None else '' for x in next(rows)]
    idx = {h: i for i, h in enumerate(header)}
    for need in ('변대주번호', '변대주명', 'DCU ID', '인입망 통신방식', '회선상태'):
        if need not in idx:
            raise SystemExit(f'대장 시트에 {need!r} 컬럼이 없다 — 헤더: {header}')

    ledger = {}
    dupes = set()
    for r in rows:
        if not r or all(x is None for x in r):
            continue
        pole_no = unformula(r[idx['변대주번호']])
        dcuid = unformula(r[idx['DCU ID']])
        if not pole_no or not dcuid:
            continue
        rec = {
            'DCUID': dcuid,
            '변대주명': unformula(r[idx['변대주명']]),
            '통신방식': unformula(r[idx['인입망 통신방식']]),
            '회선상태': unformula(r[idx['회선상태']]),
        }
        if pole_no in ledger and ledger[pole_no]['DCUID'] != dcuid:
            # 같은 전주에 DCU 가 둘 이상 = 어느 쪽인지 확정 불가 -> 부착하지 않는다
            dupes.add(pole_no)
        ledger.setdefault(pole_no, rec)
    for p in dupes:
        ledger.pop(p, None)
    return ledger, dupes


def load_removal_poles(xlsx_path):
    """'검토' 시트에서 철거 확정(검토=해지) 변대주번호 집합을 뽑는다.

    철거되는 DCU 에 물린 계기는 LTE 로 전환되므로 통신방식을 LTE 로 본다.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
    if SHEET_REVIEW not in wb.sheetnames:
        print(f'경고: {SHEET_REVIEW!r} 시트가 없어 철거예정 판정을 건너뛴다')
        return set()
    ws = wb[SHEET_REVIEW]
    rows = ws.iter_rows(min_row=2, values_only=True)
    header = [str(x).strip() if x is not None else '' for x in next(rows)]
    idx = {h: i for i, h in enumerate(header)}
    for need in ('변대주번호', '검토'):
        if need not in idx:
            raise SystemExit(f'검토 시트에 {need!r} 컬럼이 없다 — 헤더: {header}')
    poles = set()
    for r in rows:
        if not r or all(x is None for x in r):
            continue
        no = unformula(r[idx['변대주번호']])
        if no and unformula(r[idx['검토']]) == '해지':
            poles.add(no)
    return poles


def main():
    site_path = sys.argv[1] if len(sys.argv) > 1 else 'jongno-combined/data/jongno-site-data.json'
    xlsx_path = sys.argv[2] if len(sys.argv) > 2 else 'data/reference/간선망_해지_정지대상.xlsx'

    ledger, dupes = load_dcu_ledger(xlsx_path)
    print(f'DCU 대장 고유 변대주번호 {len(ledger)}건 (다중DCU 제외 {len(dupes)}건)')
    removal = load_removal_poles(xlsx_path)
    print(f'철거 확정(검토=해지) 변대주 {len(removal)}건')

    with open(site_path, encoding='utf-8') as f:
        site = json.load(f)

    stamp = datetime.now(KST).strftime('%Y%m%d-%H%M%S')
    backup = site_path.replace('.json', f'.backup-DCUID부착전-{stamp}.json')
    shutil.copy2(site_path, backup)
    print(f'백업 {backup}')

    c = Counter()
    label_mismatch = []
    for item in site:
        # 재실행 대비 — 지난 회차 값이 남아 조용히 틀리지 않게 매번 비우고 다시 채운다
        item.pop('DCUID', None)
        item.pop('통신방식', None)
        pole = str(item.get('변대주') or '').strip()
        if not pole or pole == '0':
            c['변대주없음'] += 1
            continue
        rec = ledger.get(pole)
        if not rec:
            # 대장에 DCU 가 없는 전주 = 개별 LTE 회선
            item['통신방식'] = 'LTE'
            c['대장미등재_LTE'] += 1
            continue
        # 변대주명 교차검증 — 양쪽 이름이 다르면 전산화번호가 같아도 같은 전주라 단정할 수 없다.
        # 틀린 DCU 를 현장에 보여주는 쪽이 비어 있는 것보다 위험하므로 부착하지 않는다.
        label = str(item.get('변대주라벨') or '').strip()
        if label and rec['변대주명'] and label != rec['변대주명']:
            c['라벨불일치제외'] += 1
            label_mismatch.append((pole, label, rec['변대주명']))
            continue
        item['DCUID'] = rec['DCUID']
        c['부착'] += 1
        # DCU 가 있어도 LTE 로 보는 두 경우 (영준님 지시 2026-08-13).
        # DCU ID 는 어느 전주 소속이었는지 알 수 있게 남긴다.
        #   1) 회선상태 해지·정지 — DCU 회선이 이미 끊겼거나 끊길 대상
        #   2) 검토 시트 철거 확정 — DCU 자체가 철거되어 계기가 LTE 로 전환
        if pole in removal:
            item['통신방식'] = 'LTE'
            c['철거예정_LTE'] += 1
        elif rec['회선상태'] in ('해지', '정지'):
            item['통신방식'] = 'LTE'
            c[f'해지정지_LTE_{rec["회선상태"]}'] += 1
        elif rec['통신방식']:
            item['통신방식'] = rec['통신방식']
            c[f'통신방식_{rec["통신방식"]}'] += 1

    # 원본과 같은 compact 포맷 유지 — 폰이 직접 받는 파일이라 indent 를 넣으면 용량이 커진다
    with open(site_path, 'w', encoding='utf-8') as f:
        json.dump(site, f, ensure_ascii=False)

    print(f'총 {len(site)}건 — DCU부착 {c["부착"]} / 대장미등재=LTE {c["대장미등재_LTE"]} / 변대주없음 {c["변대주없음"]}')
    print('통신방식 ' + ' / '.join(f'{k.split("_", 1)[1]} {v}' for k, v in sorted(c.items())
                                  if k.startswith('통신방식_')) +
          f' / LTE {c["대장미등재_LTE"]}')
    print('  ㄴ그중 DCU 있으나 LTE 처리 — 철거확정 '
          f'{c["철거예정_LTE"]} / 회선해지 {c["해지정지_LTE_해지"]} · 회선정지 {c["해지정지_LTE_정지"]}')
    print(f'변대주라벨 vs 대장 변대주명 불일치로 제외 {c["라벨불일치제외"]}건'
          f' (고유 전주 {len({m[0] for m in label_mismatch})}개)')
    for m in sorted({m for m in label_mismatch}):
        print(f'  불일치 변대주={m[0]} 종로라벨={m[1]!r} 대장명={m[2]!r}')

    # 부착값 자기검증 — DCUID 는 변대주로 시작하는 10자리여야 한다
    bad = [i for i in site if i.get('DCUID')
           and not (len(i['DCUID']) == 10 and i['DCUID'].startswith(str(i.get('변대주') or '')))]
    print(f'자기검증 이상 {len(bad)}건' + (f' 예시 {bad[0].get("변대주")}/{bad[0].get("DCUID")}' if bad else ''))


if __name__ == '__main__':
    main()
