"""
종로 실효 엑셀(교체 전) vs site-data.json 종로(교체 후) 매칭 검증 (v2)

매칭 키 = "법정동 번지" — 호수/층/공동주택명 모두 무시
도로명은 교차검증용 (일치 시 신뢰도 ↑)
"""

import os, re, json, importlib.util
from collections import defaultdict, Counter

# 좌표추출_v2.py 의 함수 재활용
spec = importlib.util.spec_from_file_location("zv2", os.path.join(os.path.dirname(__file__), '..', '좌표추출_v2.py'))
zv2 = importlib.util.module_from_spec(spec)
spec.loader.exec_module(zv2)

import openpyxl

EXCEL     = '/Users/woodelight/Library/Mobile Documents/com~apple~CloudDocs/종로 실효리스트.xlsx'
SITE_DATA = '/Users/woodelight/Projects/ami-work/data/site-data.json'
OUT_JSON  = '/Users/woodelight/Projects/ami-work/research/jongno_overlap.json'

SINGLE = {'17', '19', '25', '26', '27', '53'}
THREE  = {'45', '46', '47', '55'}

def phase_of(meter):
    s = zv2.fix_meter_no(meter)
    code = s[2:4] if len(s) >= 4 else ''
    if code in SINGLE: return '단상'
    if code in THREE:  return '삼상'
    return '미파악'

def norm_jibun(dong, beonji):
    """엑셀: 법정동 + 번지 → '가회동 31-11'"""
    if not dong or not beonji: return None
    d = re.sub(r'\s+', '', str(dong))
    b = re.sub(r'\s+', '', str(beonji))
    if not d or not b: return None
    return f'{d} {b}'

def jibun_from_site_addr(addr):
    """site-data 주소 → '가회동 31-11'"""
    if not addr: return None
    s = str(addr).strip()
    s = re.sub(r'^서울특별시\s*', '', s)
    s = re.sub(r'^서울\s*', '', s)
    s = re.sub(r'^종로구\s*', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    # "동/가/로N가" + 번지(산N-N 포함)
    m = re.match(r'^(\S+(?:동|가))\s+(산?\d+(?:-\d+)?)', s)
    if m: return f'{m.group(1)} {m.group(2)}'
    return s  # 패턴 못 잡으면 통째로

def norm_road(addr):
    if not addr: return ''
    s = re.sub(r'^서울특별시\s*', '', str(addr))
    s = re.sub(r'^서울\s*', '', s)
    s = re.sub(r'^종로구\s*', '', s)
    s = re.sub(r'\s+\d+호\s*$', '', s)         # 끝의 "102호" 제거
    s = re.sub(r'\s+\d+층\s*$', '', s)         # 끝의 "3층" 제거
    s = re.sub(r'(\d+(?:-\d+)?)\s*\d+호\s*$', r'\1', s)  # "22-4302호" → "22-4"
    return zv2.trim_addr(s)


# ── 엑셀 로드 ────────────────────────────────────────────────
print('[1] 엑셀 로드 중...')
wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb['저압']

excel_by_jibun = defaultdict(list)
no_dong_beonji = 0
empty_row      = 0
beonji_recovered = 0
for r in range(2, ws.max_row + 1):
    meter = ws.cell(r, 26).value
    dong  = ws.cell(r, 6).value
    beonji= ws.cell(r, 7).value
    road  = ws.cell(r, 13).value
    if not meter:
        empty_row += 1
        continue
    # 번지 폴백: 컬럼 5(지번주소1)에서 추출
    if dong and not beonji:
        jibun5 = ws.cell(r, 5).value
        if jibun5:
            s = re.sub(rf'^{re.escape(str(dong).strip())}\s*', '', str(jibun5).strip())
            m = re.match(r'^(산?\d+(?:-\d+)?)', s)
            if m:
                beonji = m.group(1)
                beonji_recovered += 1
    key = norm_jibun(dong, beonji)
    if not key:
        no_dong_beonji += 1
        continue
    excel_by_jibun[key].append({
        'meter': zv2.fix_meter_no(meter),
        'phase': phase_of(meter),
        'road': norm_road(road) if road else '',
        'raw_road': str(road).strip() if road else '',
        'raw_jibun': f'{dong} {beonji}',
    })

print(f'    엑셀 행: {ws.max_row-1:,}')
print(f'      └ 빈 행 (계기없음): {empty_row:,}')
print(f'      └ 동/번지 누락    : {no_dong_beonji:,}')
print(f'      └ 번지 복구 성공  : {beonji_recovered:,}')
print(f'      └ 매칭 데이터     : {sum(len(v) for v in excel_by_jibun.values()):,}')
print(f'    고유 지번 키: {len(excel_by_jibun):,}')


# ── site-data 종로 로드 ──────────────────────────────────────
print('\n[2] site-data 종로 필터링 중...')
with open(SITE_DATA) as f:
    site = json.load(f)

site_by_jibun = defaultdict(list)
for x in site:
    blob = (x.get('주소', '') + ' ' + x.get('도로명주소', ''))
    if '종로' not in blob: continue
    key = jibun_from_site_addr(x.get('주소', ''))
    if not key: continue
    site_by_jibun[key].append({
        'meter': x.get('계기번호', ''),
        'phase': phase_of(x.get('계기번호', '')),
        'road': norm_road(x.get('도로명주소', '')),
        'raw_addr': x.get('주소', ''),
        'raw_road': x.get('도로명주소', ''),
        'sangho': x.get('상호', ''),
    })

print(f'    site-data 종로 고유 지번: {len(site_by_jibun):,}')


# ── 매칭 ─────────────────────────────────────────────────────
common = set(excel_by_jibun) & set(site_by_jibun)
excel_only = set(excel_by_jibun) - set(site_by_jibun)
site_only  = set(site_by_jibun) - set(excel_by_jibun)

# 신뢰도 등급:
#   green : 개수+phase+도로명 일치
#   yellow: 개수+phase 일치, 도로명 깨짐
#   orange: 개수만 일치, phase 불일치
#   red   : 개수 불일치
green, yellow, orange, red = [], [], [], []

for key in common:
    e = excel_by_jibun[key]
    s = site_by_jibun[key]
    if len(e) != len(s):
        red.append((key, len(e), len(s)))
        continue
    ep = Counter(m['phase'] for m in e)
    sp = Counter(m['phase'] for m in s)
    if ep != sp:
        orange.append((key, len(e), dict(ep), dict(sp)))
        continue
    # 도로명 교차검증 — 엑셀과 site 도로명 정규화 후 비교
    e_roads = {m['road'] for m in e if m['road']}
    s_roads = {m['road'] for m in s if m['road']}
    if e_roads & s_roads:
        green.append((key, len(e), dict(ep), e[0]['raw_road'], s[0]['raw_road']))
    else:
        yellow.append((key, len(e), dict(ep), e[0]['raw_road'], s[0]['raw_road']))

excel_only_meters = sum(len(excel_by_jibun[k]) for k in excel_only)
site_only_meters  = sum(len(site_by_jibun[k])  for k in site_only)
green_meters  = sum(n for _, n, _, _, _ in green)
yellow_meters = sum(n for _, n, _, _, _ in yellow)
orange_meters = sum(n for _, n, _, _ in orange)
red_meters_excel = sum(en for _, en, _ in red)


# ── 보고 ─────────────────────────────────────────────────────
print('\n[3] 결과 (매칭 키 = 법정동+번지)')
print(f'  공통 지번 키: {len(common):,}')
print(f'    🟢 동+번지+개수+phase+도로명 모두 일치 (확실)         : 키 {len(green):,} / 계기 {green_meters:,}')
print(f'    🟡 동+번지+개수+phase 일치, 도로명 다름 (검증요)        : 키 {len(yellow):,} / 계기 {yellow_meters:,}')
print(f'    🟠 동+번지+개수 일치, phase 불일치 (false positive 가능): 키 {len(orange):,} / 계기 {orange_meters:,}')
print(f'    🔴 동+번지 같음, 개수 다름 (대형건물 등)                : 키 {len(red):,} / 엑셀 측 계기 {red_meters_excel:,}')
print(f'  엑셀만 있음 (미작업): 지번 {len(excel_only):,} / 계기 {excel_only_meters:,}')
print(f'  site만 있음 (이상)  : 지번 {len(site_only):,} / 계기 {site_only_meters:,}')

print('\n🟢 확실 (top 10):')
for key, n, ph, er, sr in sorted(green, key=lambda x: -x[1])[:10]:
    print(f'    [{n:2d}건 {ph}] {key}  엑셀도로명="{er}"  site="{sr}"')

print('\n🟡 도로명 다름 (top 10) — 진짜 같은 건물인지 봐야:')
for key, n, ph, er, sr in yellow[:10]:
    print(f'    [{n}건 {ph}] {key}  엑셀="{er}" ≠ site="{sr}"')

print('\n🟠 phase 불일치 (top 10):')
for key, n, ep, sp in orange[:10]:
    print(f'    [{n}건] {key}  엑셀{ep} ≠ site{sp}')

print('\n🔴 개수 불일치 큰 차이 (top 10):')
for key, en, sn in sorted(red, key=lambda x: -abs(x[1]-x[2]))[:10]:
    print(f'    엑셀{en:3d} vs site{sn:3d} : {key}')


# 저장
os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
with open(OUT_JSON, 'w') as f:
    json.dump({
        'match_key': '법정동+번지',
        'summary': {
            'common_keys': len(common),
            'green':  {'keys': len(green),  'meters': green_meters},
            'yellow': {'keys': len(yellow), 'meters': yellow_meters},
            'orange': {'keys': len(orange), 'meters': orange_meters},
            'red':    {'keys': len(red),    'excel_meters': red_meters_excel},
            'excel_only': {'keys': len(excel_only), 'meters': excel_only_meters},
            'site_only':  {'keys': len(site_only),  'meters': site_only_meters},
        },
        'green':  [{'key': k, 'n': n, 'phase': ph, 'excel_road': er, 'site_road': sr} for k, n, ph, er, sr in green],
        'yellow': [{'key': k, 'n': n, 'phase': ph, 'excel_road': er, 'site_road': sr} for k, n, ph, er, sr in yellow],
        'orange': [{'key': k, 'n': n, 'excel_phase': ep, 'site_phase': sp} for k, n, ep, sp in orange],
        'red':    [{'key': k, 'excel': en, 'site': sn} for k, en, sn in red],
    }, f, ensure_ascii=False, indent=2)
print(f'\n→ 저장: {OUT_JSON}')
